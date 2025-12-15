import pdfplumber

def extract_purchase_items_from_pdf(file_path):
    items = []

    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue

            for row in table:
                # Skip header rows
                if "Item Description" in row[1]:
                    continue

                try:
                    code = row[1]
                    name = row[2]
                    qty = row[5]
                    pack = row[6]
                    batch = row[7]
                    expiry = row[9]
                    mrp = row[10]
                    cost = row[12]
                    net = row[15]

                    items.append({
                        "product_code": code,
                        "name": name,
                        "qty": float(qty or 0),
                        "pack": pack,
                        "batch_no": batch,
                        "expiry": expiry,
                        "mrp": float(mrp or 0),
                        "cost": float(cost or 0),
                        "net_value": float(net or 0),
                    })

                except:
                    continue

    return items


import openpyxl
from decimal import Decimal
import re
import csv
import os

def extract_items_from_csv(file_content_or_path):
    """
    Extract items from CSV file.
    Accepts either a file path (str) or file content (bytes/str) for in-memory processing.
    """
    items = []

    # Handle both file path (string) and file content (bytes/str/io object)
    if isinstance(file_content_or_path, str) and os.path.exists(file_content_or_path):
        # Legacy: file path provided
        try:
            with open(file_content_or_path, "r", encoding="utf-8") as f:
                lines = f.read()
        except UnicodeDecodeError:
            try:
                with open(file_content_or_path, "r", encoding="latin-1") as f:
                    lines = f.read()
            except UnicodeDecodeError:
                with open(file_content_or_path, "r", encoding="utf-8", errors="ignore") as f:
                    lines = f.read()
    else:
        # In-memory processing: file_content_or_path is bytes or string
        if isinstance(file_content_or_path, bytes):
            # Try UTF-8 first
            try:
                lines = file_content_or_path.decode("utf-8")
            except UnicodeDecodeError:
                # Fallback to latin-1
                try:
                    lines = file_content_or_path.decode("latin-1")
                except UnicodeDecodeError:
                    # Last resort: ignore errors
                    lines = file_content_or_path.decode("utf-8", errors="ignore")
        else:
            # Already a string
            lines = file_content_or_path

    # Normalize spacing → convert multiple spaces/tabs into single comma
    normalized = ",".join(
        part for part in lines.replace("\t", " ").split(" ") if part.strip() != ""
    )

    rows = normalized.split("\n")

    if len(rows) < 2:
        return []

    # Parse header
    header = rows[0].split(",")

    # Build index map safely (case-insensitive)
    idx = {}
    idx_lower = {}
    for i, col in enumerate(header):
        col_clean = col.strip()
        idx[col_clean] = i
        idx_lower[col_clean.lower()] = i

    # Flexible column name matching - try multiple variations
    def find_column_index(possible_names):
        """Find column index by trying multiple possible names (case-insensitive)"""
        for name in possible_names:
            # Try exact match first
            if name in idx:
                return idx[name]
            # Try case-insensitive match
            name_lower = name.lower()
            if name_lower in idx_lower:
                return idx_lower[name_lower]
            # Try partial match (contains)
            for col_name, col_idx in idx_lower.items():
                if name_lower in col_name or col_name in name_lower:
                    return col_idx
        return None

    # Find required columns with flexible matching
    name_idx = find_column_index([
        "ItemName", "itemname", "item_name", "Item Name", "Item", "Product Name", 
        "ProductName", "product_name", "Name", "Medicine Name", "MedicineName",
        "medicine name", "medicine_name", "Product", "product"
    ])
    qty_idx = find_column_index([
        "InvQty", "invqty", "inv_qty", "Inv Qty", "Quantity", "Qty", "qty", 
        "QTY", "quantity", "Qty Pack", "QtyPack", "qty_pack", "qty pack",
        "Qty Packs", "qty_packs", "QtyPacks"
    ])
    rate_idx = find_column_index([
        "SaleRate", "salerate", "sale_rate", "Sale Rate", "Rate", "rate", 
        "RATE", "Price", "price", "Unit Price", "UnitPrice", "unit_price",
        "Cost", "cost", "Unit Cost", "UnitCost", "unit_cost", "Sale Price",
        "sale_price", "SalePrice"
    ])

    # Check if we found required columns
    if name_idx is None or qty_idx is None or rate_idx is None:
        # Try to provide helpful error message
        available_cols = ', '.join([col.strip() for col in header])
        missing = []
        if name_idx is None:
            missing.append("item name")
        if qty_idx is None:
            missing.append("quantity")
        if rate_idx is None:
            missing.append("rate/price")
        
        print(f"Missing required columns: {', '.join(missing)}")
        print(f"Available columns: {available_cols}")
        return []   # cannot parse this CSV

    items = []
    for row in rows[1:]:
        if not row.strip():
            continue

        parts = row.split(",")

        try:
            # Get values safely using the found indices
            name = parts[name_idx].strip().strip('"').strip("'") if name_idx < len(parts) else ""
            qty = parts[qty_idx].strip().strip('"').strip("'") if qty_idx < len(parts) else "0"
            rate = parts[rate_idx].strip().strip('"').strip("'") if rate_idx < len(parts) else "0"
        except (IndexError, AttributeError) as e:
            continue

        if name == "":
            continue

        items.append({
            "product_code": "",
            "name": name,
            "qty": qty,
            "rate": rate,
            "net_value": "0",
        })

    return items


def extract_items_from_excel(file_content_or_path):
    """
    Extract purchase items from Excel (.xlsx)
    Expected same columns as CSV.
    Accepts either a file path (str) or file-like object (BytesIO) for in-memory processing.
    """
    # openpyxl.load_workbook can accept both file paths and file-like objects
    wb = openpyxl.load_workbook(file_content_or_path, data_only=True)
    ws = wb.active

    # Read header row
    headers = [str(cell.value).strip().lower() for cell in ws[1]]

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = dict(zip(headers, row))

        items.append({
            "product_code": values.get("product_code", ""),
            "name": values.get("name", ""),
            "qty": values.get("qty", "0"),
            "rate": values.get("rate", "0"),
            "net_value": values.get("net_value", "0"),
        })

    return items
