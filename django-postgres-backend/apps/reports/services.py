import io
import uuid
from datetime import date
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from django.utils import timezone

from apps.sales.models import SalesInvoice
from apps.compliance.models import H1RegisterEntry, NDPSDailyEntry
from apps.inventory.models import InventoryMovement
from apps.settingsx.services import get_setting
from apps.inventory.services import near_expiry
from apps.catalog.models import Product

REPORT_UI_NAMES = {
    "SALES_REGISTER": "Sales_Report",
    "H1_REGISTER": "H1_Register",
    "NDPS_DAILY": "NDPS_Daily",
    "STOCK_LEDGER": "Purchase_Report",
    "EXPIRY_STATUS": "Expiry_Report",
    "TOP_SELLING": "Top_Selling_Report",
}



def _auto_width(ws):
    """Auto adjust column width"""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_len + 2


def generate_report_file(export):
    """
    Generates an Excel file completely IN MEMORY
    and returns (filename, BytesIO buffer)
    """

    today = date.today().strftime("%Y-%m-%d")
    short_uid = uuid.uuid4().hex[:6]

    ui_name = REPORT_UI_NAMES.get(export.report_type, export.report_type)
    filename = f"{ui_name}_{today}_{short_uid}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = export.report_type

    params = export.params or {}

    # ------------------------------
    # SALES REGISTER
    # ------------------------------
    if export.report_type == "SALES_REGISTER":
        ws.append([
            "Invoice No", "Invoice Date", "Customer", "Product", "Batch",
            "Qty", "Rate", "Tax %", "Tax Amt", "Line Total", "Net Total"
        ])

        qs = SalesInvoice.objects.prefetch_related("lines__product", "lines__batch_lot")

        if params.get("date_from"):
            qs = qs.filter(invoice_date__date__gte=params["date_from"])
        if params.get("date_to"):
            qs = qs.filter(invoice_date__date__lte=params["date_to"])
        if params.get("customer"):
            qs = qs.filter(customer_id=params["customer"])
        if params.get("location"):
            qs = qs.filter(location_id=params["location"])

        for inv in qs:
            for line in inv.lines.all():
                ws.append([
                    inv.invoice_no,
                    inv.invoice_date.strftime("%Y-%m-%d"),
                    inv.customer.name if inv.customer else "",
                    line.product.name,
                    line.batch_lot.batch_no,
                    float(line.qty_base),
                    float(line.rate_per_base),
                    float(line.tax_percent),
                    float(line.tax_amount),
                    float(line.line_total),
                    float(inv.net_total)
                ])

    # ------------------------------
    # H1 REGISTER
    # ------------------------------
    elif export.report_type == "H1_REGISTER":
        ws.append([
            "Invoice No", "Entry Date", "Product", "Batch", "Qty",
            "Patient", "Doctor", "Doctor Reg No"
        ])

        qs = H1RegisterEntry.objects.select_related("invoice", "product", "batch_lot")

        if params.get("date_from"):
            qs = qs.filter(entry_date__date__gte=params["date_from"])
        if params.get("date_to"):
            qs = qs.filter(entry_date__date__lte=params["date_to"])
        if params.get("invoice"):
            qs = qs.filter(invoice_id=params["invoice"])
        if params.get("product"):
            qs = qs.filter(product_id=params["product"])

        for e in qs:
            ws.append([
                e.invoice.invoice_no if e.invoice else "",
                e.entry_date.strftime("%Y-%m-%d"),
                e.product.name if e.product else "",
                e.batch_lot.batch_no if e.batch_lot else "",
                float(e.qty_issued_base),
                e.patient_name,
                e.doctor_name,
                e.doctor_reg_no,
            ])

    # ------------------------------
    # NDPS DAILY
    # ------------------------------
    elif export.report_type == "NDPS_DAILY":
        ws.append(["Date", "Product", "Opening", "Issued", "Closing"])

        qs = NDPSDailyEntry.objects.select_related("product")

        if params.get("date_from"):
            qs = qs.filter(date__gte=params["date_from"])
        if params.get("date_to"):
            qs = qs.filter(date__lte=params["date_to"])
        if params.get("product"):
            qs = qs.filter(product_id=params["product"])

        for e in qs:
            ws.append([
                e.date.strftime("%Y-%m-%d"),
                e.product.name,
                float(e.opening_qty_base),
                float(e.out_qty_base),
                float(e.closing_qty_base),
            ])

    # ------------------------------
    # STOCK LEDGER
    # ------------------------------
    elif export.report_type == "STOCK_LEDGER":
        ws.append(["Movement Date", "Location", "Product", "Batch", "Reason", "Qty Change"])

        qs = InventoryMovement.objects.select_related("location", "batch_lot", "batch_lot__product")

        if params.get("date_from"):
            qs = qs.filter(created_at__date__gte=params["date_from"])
        if params.get("date_to"):
            qs = qs.filter(created_at__date__lte=params["date_to"])
        if params.get("location"):
            qs = qs.filter(location_id=params["location"])

        for move in qs:
            ws.append([
                move.created_at.date().strftime("%Y-%m-%d"),
                move.location.name,
                move.batch_lot.product.name,
                move.batch_lot.batch_no,
                move.reason,
                float(move.qty_change_base),
            ])

    # ------------------------------
    # EXPIRY STATUS REPORT
    # ------------------------------
    elif export.report_type == "EXPIRY_STATUS":
        ws.append([
            "Medicine Name", "Batch Number", "Category", "Quantity",
            "Stock Value", "Expiry Date", "Days Left", "Status"
        ])

        warn_days = int(get_setting("ALERT_EXPIRY_WARNING_DAYS", "60") or 60)
        crit_days = int(get_setting("ALERT_EXPIRY_CRITICAL_DAYS", "30") or 30)

        rows = near_expiry(
            days=warn_days,
            location_id=params.get("location")
        )

        products = {p.id: p for p in Product.objects.filter(id__in=[r["product_id"] for r in rows])}

        today = date.today()

        for r in rows:
            exp = r.get("expiry_date")
            days_left = (exp - today).days if exp else None
            prod = products.get(r.get("product_id"))

            if days_left is None:
                status_txt = "Unknown"
            elif days_left <= crit_days:
                status_txt = "Critical"
            elif days_left <= warn_days:
                status_txt = "Warning"
            else:
                status_txt = "Safe"

            price_per_base = 0
            if prod and prod.units_per_pack:
                try:
                    price_per_base = float(prod.mrp) / float(prod.units_per_pack)
                except:
                    price_per_base = 0

            stock_value = round(float(r.get("stock_base", 0)) * price_per_base, 2)

            ws.append([
                getattr(prod, "name", ""),
                r.get("batch_no"),
                getattr(getattr(prod, "category", None), "name", ""),
                r.get("stock_base"),
                stock_value,
                exp.strftime("%Y-%m-%d") if exp else "",
                days_left,
                status_txt,
            ])

    # Auto-size columns
    _auto_width(ws)

    # ------------------------------
    # GENERATE FILE IN MEMORY ONLY
    # ------------------------------
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return filename, buffer